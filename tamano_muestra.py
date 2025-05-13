import requests
from bs4 import BeautifulSoup
from Bio import Entrez
import re
import os
import pandas as pd
import time
import concurrent.futures
import logging
import pycountry
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import transformers
from itertools import count
from transformers import pipeline
from sentence_transformers import SentenceTransformer
import numpy as np
import torch
from sklearn.metrics.pairwise import cosine_similarity
from typing import List, Dict, Optional, Tuple

# Configuración inicial
DEBUG_DIR = "debug_articles"
os.makedirs(DEBUG_DIR, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('extraction_log.log')
    ]
)
transformers.logging.set_verbosity_error()
article_counter = count(1)
load_dotenv()

# Constantes
DEFAULT_RETURN = "No disponible"
MAX_RETRIES = 3
INITIAL_WAIT_TIME = 6
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
MAX_WORKERS = 4  # Número máximo de hilos para concurrent.futures
QA_CONFIDENCE_THRESHOLD = 0.7

# Configuración de la API de Entrez
Entrez.email = os.getenv('EMAIL')
Entrez.api_key = os.getenv('API_KEY')



class RAGSystem:
    """Sistema de Retrieval-Augmented Generation para análisis de artículos científicos."""
    
    def __init__(self):
        self.model = SentenceTransformer("all-MiniLM-L12-v2", device="cpu")
        self.knowledge_base = []
        self.embeddings = None
        
    def add_to_knowledge_base(self, documents: List[Dict]) -> None:
        texts = [self._format_document(doc) for doc in documents]
        self.knowledge_base.extend(documents)
        
        new_embeddings = self.model.encode(texts, convert_to_tensor=True)
        if self.embeddings is None:
            self.embeddings = new_embeddings
        else:
            self.embeddings = torch.cat([self.embeddings, new_embeddings])
    
    def _format_document(self, doc: Dict) -> str:
        return f"Título: {doc.get('Titulo', '')}\nAutores: {doc.get('Autores', '')}\nResumen: {doc.get('Abstract', '')}\nTexto: {doc.get('Texto_Completo', '')}"
    
    def _split_into_chunks(self, text: str, chunk_size: int = 500) -> List[str]:
        words = text.split()
        return [' '.join(words[i:i + chunk_size]) for i in range(0, len(words), chunk_size)]
    
    def query(self, question: str, context: str, top_k: int = 3) -> List[Dict]:
        chunks = self._split_into_chunks(context)
        chunk_embeddings = self.model.encode(chunks, convert_to_tensor=True)
        
        question_embedding = self.model.encode(question, convert_to_tensor=True)
        similarities = cosine_similarity(
            question_embedding.unsqueeze(0).cpu().numpy(),
            chunk_embeddings.cpu().numpy()
        )[0]
        
        top_indices = np.argsort(similarities)[-top_k:][::-1]
        return [{"text": chunks[i], "score": float(similarities[i])} for i in top_indices]

def init_qa_model():
    return pipeline(
        "question-answering",
        model="ktrapeznikov/biobert_v1.1_pubmed_squad_v2",
        tokenizer="ktrapeznikov/biobert_v1.1_pubmed_squad_v2",
        device=-1
    )

def validate_sample_size(number: str) -> bool:
    """Valida si un número extraído es un tamaño de muestra plausible."""
    try:
        n = int(number)
        # Rangos plausibles para estudios clínicos/microbiológicos
        return 10 <= n <= 1000000  # Ajustar según el dominio específico
    except ValueError:
        return False

def extract_numbers_with_context(text: str, patterns: List[str]) -> List[Tuple[str, str]]:
    """Extrae números con su contexto cercano para análisis semántico."""
    results = []
    for pattern in patterns:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            context_start = max(0, match.start() - 50)
            context_end = min(len(text), match.end() + 50)
            context = text[context_start:context_end].replace('\n', ' ')
            results.append((match.group(1), context))
    return results

def filter_sample_size_candidates(candidates: List[Tuple[str, str]], qa_pipeline) -> Optional[str]:
    """Filtra candidatos usando un modelo de QA para identificar el verdadero tamaño de muestra."""
    question = "Is this number the exact total sample size of the study? Answer yes or no."
    
    for number, context in candidates:
        try:
            result = qa_pipeline(question=question, context=context)
            if result['answer'].lower().startswith('yes') and result['score'] > QA_CONFIDENCE_THRESHOLD:
                return number.replace(',', '')
        except:
            continue
    return None

def enhanced_sample_size_extraction(text: str, qa_pipeline, rag_system: RAGSystem) -> str:
    # 1. Extracción con regex mejorada
    patterns = [
        r"(?:n\s*[=:]\s*|sample\s*size\s*of|included\s*)(\d{1,3}(?:,\d{3})*)",
        r"(?:total\s+(?:of\s+)?|participants[:]?\s*)(\d{1,3}(?:,\d{3})*)\s+(?:participants|patients|women|subjects)",
        r"\b(\d{2,})\s*(?:subjects|samples|cases|individuals|women|patients)\b",
        r"study\s*population\s*of\s*(\d+)",
        r"\b(\d+)\s*participants\b",
        r"total\s*number\s*of\s*(\d+)\s*(?:cases|samples)"
    ]
    
    # Extraer números con contexto
    candidates = extract_numbers_with_context(text, patterns)
    
    # Filtrar solo números válidos
    valid_candidates = [(num, ctx) for num, ctx in candidates if validate_sample_size(num.replace(',', ''))]
    
    if valid_candidates:
        # 2. Filtrado semántico con BioBERT
        best_candidate = filter_sample_size_candidates(valid_candidates, qa_pipeline)
        if best_candidate:
            return best_candidate
        
        # 3. Fallback: Seleccionar el número más grande en métodos/resultados
        numbers = [int(num.replace(',', '')) for num, _ in valid_candidates]
        return str(max(numbers)) if numbers else DEFAULT_RETURN
    
    # 4. Búsqueda con RAG como último recurso
    try:
        rag_results = rag_system.query("What is the total sample size of the study? Report only the number.", text, top_k=3)
        for result in rag_results:
            numbers = re.findall(r'\d{2,}', result['text'])
            valid_numbers = [n for n in numbers if validate_sample_size(n)]
            if valid_numbers:
                return max(valid_numbers, key=int)
    except Exception as e:
        logging.error(f"Error en RAG: {e}")
    
    return DEFAULT_RETURN

def search_pubmed_mesh(term_mesh: str, max_results: int = 100) -> List[str]:
    term_mesh += " AND free full text[Filter]"
    handle = Entrez.esearch(db="pubmed", term=term_mesh, retmax=max_results)
    search_results = Entrez.read(handle)
    handle.close()
    return search_results['IdList']

def search_pubmed_mesh(term_mesh: str, max_results: int = 100) -> List[str]:
    term_mesh += " AND free full text[Filter]"
    handle = Entrez.esearch(db="pubmed", term=term_mesh, retmax=max_results)
    search_results = Entrez.read(handle)
    handle.close()
    return search_results['IdList']

def get_country_from_affiliation(soup: BeautifulSoup) -> str:
    affiliations_section = soup.find('div', class_='affiliations')
    if not affiliations_section:
        return DEFAULT_RETURN
    
    affiliation_text = affiliations_section.get_text(separator=" ")
    countries = {country.name.lower(): country.name for country in pycountry.countries}
    found_countries = set()
    
    for country_name in countries:
        if country_name in affiliation_text.lower():
            found_countries.add(countries[country_name])
    
    return ", ".join(found_countries) if found_countries else DEFAULT_RETURN

def get_pmc_full_text(pmcid_url: str) -> str:
    """Extrae el texto completo de un artículo en PMC."""
    try:
        response = requests.get(pmcid_url, headers={'User-Agent': USER_AGENT}, timeout=10)
        if response.status_code == 200:
            pmc_soup = BeautifulSoup(response.text, 'html.parser')
            
            # Eliminar elementos no deseados
            for element in pmc_soup.find_all(['div', 'section']):
                if 'ref-list' in element.get('class', []):
                    element.decompose()
                if 'fig' in element.get('class', []):
                    element.decompose()
            
            # Extraer secciones principales
            sections = []
            for sec in pmc_soup.find_all(['div', 'section']):
                if sec.get('id') in ['abstract', 'methods', 'results', 'discussion']:
                    sections.append(sec.get_text(separator=' ', strip=True))
            
            return ' '.join(sections) if sections else DEFAULT_RETURN
    except Exception as e:
        logging.error(f"Error al obtener texto completo de PMC: {e}")
    
    return DEFAULT_RETURN

def extract_data(pubmed_id: str) -> Optional[Dict]:
    base_url = f"https://pubmed.ncbi.nlm.nih.gov/{pubmed_id}/"
    wait_time = INITIAL_WAIT_TIME
    
    for attempt in range(MAX_RETRIES):
        try:
            response = requests.get(base_url, headers={'User-Agent': USER_AGENT})
            
            if response.status_code == 200:
                break
            elif response.status_code in [429, 403]:
                time.sleep(wait_time)
                wait_time *= 2
            else:
                return None
        except Exception as e:
            logging.error(f"Error al acceder a {pubmed_id}: {e}")
            if attempt == MAX_RETRIES - 1:
                return None
            time.sleep(wait_time)
            wait_time *= 2
    
    if response.status_code != 200:
        return None
    
    soup = BeautifulSoup(response.text, 'html.parser')
    
    # Extraer metadatos básicos
    title = soup.find('h1', class_='heading-title').get_text(strip=True) if soup.find('h1', class_='heading-title') else DEFAULT_RETURN
    authors_section = soup.find('div', class_='authors-list')
    authors = re.findall(r"[A-ZÁÉÍÓÚÜÑ][a-záéíóúüñäëïöüßøÅåÄÖÜçÇŠšŽž]+(?:\s[A-ZÁÉÍÓÚÜÑa-záéíóúüñäëïöüßøÅåÄÖÜçÇŠšŽž\-]+)+", authors_section.get_text()) if authors_section else [] 
    
    # Extraer PMCID y verificar acceso libre
    free_access = bool(soup.find('span', class_='text', string=re.compile('Free PMC article', re.IGNORECASE)))
    pmcid_link = soup.find('a', href=re.compile(r'https://www.ncbi.nlm.nih.gov/pmc/articles/PMC\d+'))
    pmcid_url = pmcid_link['href'] if pmcid_link else "No disponible"
    
    # obtener el enlance de PMCID
    pmcid_link = soup.find('a', class_='id-link', attrs={'href': re.compile(r'pmc\.ncbi\.nlm\.nih\.gov/articles/PMC\d+')})
    pmcid_url = pmcid_link['href'] if pmcid_link else "No disponible"

    # Extraer texto completo si está disponible en PMC
    full_text = DEFAULT_RETURN
    if free_access and pmcid_url != DEFAULT_RETURN:
        full_text = get_pmc_full_text(pmcid_url)
    
    # Extraer abstract
    abstract = soup.find('div', class_='abstract-content')
    abstract_text = abstract.get_text() if abstract else ""
        
    # Extraer otros metadatos
    doi_match = re.search(r"10\.\d{4,9}/[-._;()/:A-Za-z0-9]+", soup.get_text())
    doi = doi_match.group(0) if doi_match else DEFAULT_RETURN
    
    journal_section = soup.find('button', class_='journal-actions-trigger')
    journal = journal_section.get_text(strip=True) if journal_section else DEFAULT_RETURN
    
    year_match = re.search(r"\b(19|20)\d{2}\b", soup.get_text())
    year = year_match.group(0) if year_match else DEFAULT_RETURN

    free_access = bool(soup.find('span', class_='text', string=re.compile('Free PMC article', re.IGNORECASE)))
    
    return {
        'Titulo': title,
        'PMC Free Access': "Sí" if free_access else "No",
        'Autores': ", ".join(authors),
        'Paises Afiliados': get_country_from_affiliation(soup),
        'DOI': doi,
        'NCBI ID': pubmed_id,
        'Journal': journal,
        'Año de Publicacion': year,
        'Enlace al articulo': base_url,
        'Texto_Completo': full_text,
        'Abstract': abstract_text,
        'PMCID URL': pmcid_url
    }

def process_article(pubmed_id: str, qa_pipeline, rag_system: RAGSystem) -> Tuple[str, Optional[Dict]]:
    try:
        article_data = extract_data(pubmed_id)
        if not article_data:
            return (pubmed_id, None)
        
        # Extraer tamaño de muestra con el sistema mejorado
        context = f"{article_data.get('Abstract', '')} {article_data.get('Texto_Completo', '')}"
        article_data['Numero de Participantes'] = enhanced_sample_size_extraction(
            context, 
            qa_pipeline, 
            rag_system
        )
        
        # Añadir a la base de conocimiento RAG si tiene texto completo
        if article_data['PMC Free Access'] == "Sí":
            rag_system.add_to_knowledge_base([article_data])
        
        return (pubmed_id, article_data)
    except Exception as e:
        logging.error(f"Error procesando artículo {pubmed_id}: {e}")
        return (pubmed_id, None)

if __name__ == "__main__":
    start_time = time.time()
    
    # Inicializar modelos
    logging.info("Inicializando modelos de NLP...")
    qa_pipeline = init_qa_model()
    rag_system = RAGSystem()
    
    # Búsqueda en PubMed
    term_mesh = '((((all[sb] NOT(animals [mh] NOT humans [mh])) AND (microbiota [mh])) AND (female genitalia[MeSH Terms]) ) AND (("2010/07/01"[Date - Publication] : "2024/07/21"[Date - Publication]))) NOT (Review[Publication Type])'
    
    logging.info(f"Buscando artículos con término MeSH: {term_mesh}")
    article_ids = search_pubmed_mesh(term_mesh, max_results=100)
    logging.info(f"Encontrados {len(article_ids)} artículos potenciales")
    
    # Procesamiento concurrente
    results = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        # Preparar las tareas
        future_to_id = {
            executor.submit(process_article, pubmed_id, qa_pipeline, rag_system): pubmed_id 
            for pubmed_id in article_ids
        }
        
        # Procesar los resultados a medida que estén disponibles
        completed = 0
        for future in concurrent.futures.as_completed(future_to_id):
            pubmed_id = future_to_id[future]
            completed += 1
            try:
                processed_id, article_data = future.result()
                if article_data:
                    results.append(article_data)
                    logging.info(f"[{completed}/{len(article_ids)}] Artículo {processed_id} procesado")
                    logging.info(f"  - Título: {article_data['Titulo'][:50]}...")
                    logging.info(f"  - PMCID: {article_data['PMCID URL']}")
                    logging.info(f"  - Muestra: {article_data['Numero de Participantes']}")
            except Exception as e:
                logging.error(f"Error al procesar artículo {pubmed_id}: {e}")
    
    # Guardar resultados
    if results:
        df = pd.DataFrame(results)
        output_file = 'resultados_mejorados.xlsx'
        df.to_excel(output_file, index=False, engine='openpyxl')
        
        # Formato condicional
        wb = load_workbook(output_file)
        ws = wb.active
        red_fill = PatternFill(start_color="Ff00ff", end_color="Ff00ff", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value == DEFAULT_RETURN:
                    cell.fill = red_fill
                elif cell.column_letter == 'K' and not str(cell.value).isdigit():  # Columna de tamaño de muestra
                    cell.fill = yellow_fill
        
        wb.save(output_file)
        logging.info(f"Resultados guardados en {output_file}")
        logging.info(f"Total de artículos procesados: {len(results)}")
        logging.info(f"Artículos con texto completo: {len([r for r in results if r['PMC Free Access'] == 'Sí'])}")
        logging.info(f"Tamaños de muestra encontrados: {len([r for r in results if r['Numero de Participantes'] != DEFAULT_RETURN])}")
    else:
        logging.warning("No se encontraron artículos válidos para procesar")
    
    # Estadísticas finales
    elapsed_time = time.time() - start_time
    logging.info(f"Proceso completado en {elapsed_time:.2f} segundos")
    logging.info(f"Velocidad: {len(results)/elapsed_time:.2f} artículos/segundo")
